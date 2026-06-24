"""
app.py  --  Port of Batangas (POB) Collection & Import Analytics Dashboard.

Server-backed edition: the raw BOC masterlist .xlsx is parsed ONCE on upload
(openpyxl), cleaned + flag-derived in pure Python, and held in memory as compact
records sorted by date. The browser then drives a live dashboard via JSON API
endpoints — a date-range calendar + oil/port filters that re-aggregate every tab
on demand, and click-through drill-downs (importer->products, commodity->importers,
…) each able to reveal the underlying ENTRY numbers.

Counts are by DISTINCT entry number (`ENTRY_CODE`) = one declaration, NOT by line
item (a declaration can have hundreds of line items).

Run:
    py app.py            # http://127.0.0.1:8060
"""

import base64
import bisect
import csv
import gzip
import hmac
import io
import json
import os
import re
import sys
import threading
import traceback
import urllib.request
from datetime import datetime, date, timedelta

from flask import Flask, jsonify, make_response, render_template, request

import memo_builder

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024

# ── HTTP Basic Auth (whole app is gated — repo is public, so credentials are
# read from env vars on the host, falling back to the agreed defaults) ──────
AUTH_USER = os.environ.get("DASH_USER", "admin")
AUTH_PASS = os.environ.get("DASH_PASS", "Admin@5050")


@app.before_request
def _require_login():
    auth = request.authorization
    ok = (auth and hmac.compare_digest(auth.username or "", AUTH_USER)
          and hmac.compare_digest(auth.password or "", AUTH_PASS))
    if not ok:
        return make_response(
            "Login required", 401,
            {"WWW-Authenticate": 'Basic realm="POB Assessment Dashboard"'})

# ── Supabase Storage — persist uploaded files across cold starts ──────────────
# Set SUPABASE_URL and SUPABASE_KEY (service-role key) as Render env vars.
# SUPABASE_BUCKET defaults to "masterlist". If vars are unset the app works
# normally, just without cross-restart persistence.
_SB_URL    = os.environ.get("SUPABASE_URL", "").rstrip("/")
_SB_KEY    = os.environ.get("SUPABASE_KEY", "")
_SB_BUCKET = os.environ.get("SUPABASE_BUCKET", "masterlist")

_SB_SNAP = "cache/snapshot.json.gz"   # pre-parsed records; stays under 50 MB free limit
_SB_META = "meta/upload_meta.json"   # tiny existence marker written after snapshot succeeds

_SB_AUTOLOAD_DONE = False
_SB_AUTOLOAD_LOCK = threading.Lock()


def _sb_ok():
    return bool(_SB_URL and _SB_KEY)


def _sb_upload(key, data, ctype="application/octet-stream"):
    """Upload to Supabase Storage. Returns True on success, False on failure."""
    url = f"{_SB_URL}/storage/v1/object/{_SB_BUCKET}/{key}"
    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("Authorization", f"Bearer {_SB_KEY}")
    req.add_header("Content-Type", ctype)
    req.add_header("x-upsert", "true")
    try:
        with urllib.request.urlopen(req, timeout=180):
            pass
        print(f"[supabase] saved {key} ({len(data)/1e6:.1f} MB)", flush=True)
        return True
    except Exception as e:
        print(f"[supabase] upload {key} failed: {e}", flush=True)
        return False


def _sb_download(key, on_progress=None):
    """Download from Supabase Storage. Returns bytes or None if missing."""
    url = f"{_SB_URL}/storage/v1/object/{_SB_BUCKET}/{key}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("Authorization", f"Bearer {_SB_KEY}")
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            total = int(r.headers.get("Content-Length") or 0)
            buf, done = [], 0
            while True:
                chunk = r.read(131072)  # 128 KB chunks
                if not chunk:
                    break
                buf.append(chunk)
                done += len(chunk)
                if on_progress:
                    on_progress(done, total)
            return b"".join(buf)
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print(f"[supabase] download {key} HTTP {e.code}", flush=True)
        return None
    except Exception as e:
        print(f"[supabase] download {key} failed: {e}", flush=True)
        return None


def _sb_put_meta(d):
    _sb_upload(_SB_META, json.dumps(d).encode(), "application/json")


def _sb_get_meta():
    data = _sb_download(_SB_META)
    if not data:
        return None
    try:
        return json.loads(data.decode())
    except Exception:
        return None


def _pack_ds(ds):
    """Serialize STATE["ds"] to gzip-compressed JSON (typically 5–15 MB, well under 50 MB).
    Streams records one-by-one into the gzip compressor to avoid duplicating the full
    record list in memory — peak overhead is one record at a time, not the whole set."""
    def ser_rec(r):
        row = list(r)
        row[COMPS] = list(r[COMPS])
        return row

    targets_serial = {f"{y},{m}": v for (y, m), v in ds["targets"].items()}

    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode="wb", compresslevel=6) as gz:
        # Header fields
        header = {
            "v": 1,
            "file1": ds["file1"],
            "prior_name": ds["prior_name"],
            "date_basis": ds["date_basis"],
            "dmin": ds["dmin"].isoformat(),
            "dmax": ds["dmax"].isoformat(),
            "ports": ds["ports"],
            "buckets": ds["buckets"],
            "undated": ds["undated"],
            "targets": targets_serial,
        }
        gz.write(json.dumps(header, separators=(',', ':')).encode())
        gz.write(b'\n')

        # Records streamed one line each — no full-list copy in memory
        gz.write(b'[\n')
        for i, r in enumerate(ds["records"]):
            line = json.dumps(ser_rec(r), separators=(',', ':')).encode()
            gz.write((b',' if i else b'') + line + b'\n')
        gz.write(b']\n')

        # Optional prior records
        prior_recs = ds.get("prior") and ds["prior"]["records"]
        if prior_recs:
            gz.write(b'[\n')
            for i, r in enumerate(prior_recs):
                line = json.dumps(ser_rec(r), separators=(',', ':')).encode()
                gz.write((b',' if i else b'') + line + b'\n')
            gz.write(b']\n')
        else:
            gz.write(b'null\n')

    compressed = buf.getvalue()
    print(f"[supabase] snapshot: {len(compressed)/1e6:.1f} MB compressed", flush=True)
    return compressed


def _unpack_ds(data):
    """Decompress and deserialize a snapshot blob back to a STATE["ds"] dict."""
    def de_rec(row):
        row[COMPS] = tuple(row[COMPS])
        return tuple(row)

    lines = gzip.decompress(data).split(b'\n')
    # Line 0: header JSON; line 1: start of records array '['; lines 2..N-3: record rows;
    # line N-2: ']'; line N-1: prior block ('null' or '[' … ']')
    header = json.loads(lines[0])

    # Collect records lines between the outer '[' and ']'
    recs_lines, prior_lines = [], []
    in_prior = False
    for line in lines[1:]:
        s = line.strip()
        if not s or s == b'[':
            if s == b'[' and recs_lines:
                in_prior = True
            continue
        if s == b']':
            continue
        if s == b'null':
            break
        target = prior_lines if in_prior else recs_lines
        target.append(s.lstrip(b','))

    def parse_block(raw_lines):
        result = []
        for ln in raw_lines:
            if ln:
                result.append(de_rec(json.loads(ln)))
        return result

    recs = parse_block(recs_lines)
    ords = [r[ORD] for r in recs]

    targets = {}
    for k, v in header["targets"].items():
        y, m = k.split(',')
        targets[(int(y), int(m))] = v

    prior = None
    if prior_lines:
        precs = parse_block(prior_lines)
        prior = {"records": precs, "ords": [r[ORD] for r in precs]}

    return {
        "records": recs,
        "ords": ords,
        "dmin": date.fromisoformat(header["dmin"]),
        "dmax": date.fromisoformat(header["dmax"]),
        "ports": header["ports"],
        "buckets": header["buckets"],
        "targets": targets,
        "date_basis": header["date_basis"],
        "file1": header["file1"],
        "undated": header["undated"],
        "prior": prior,
        "prior_name": header["prior_name"],
    }


def _do_load_from_supabase(_meta):
    """Background thread: download pre-parsed snapshot and restore STATE["ds"] directly.
    No Excel parsing needed — cold start goes from ~3 min down to ~30 seconds."""
    global _SB_AUTOLOAD_DONE
    try:
        def dl_progress(done, total):
            mb = done / 1e6
            tot_mb = total / 1e6 if total else 0
            stage = (f"Loading data from cloud… {mb:.0f}/{tot_mb:.0f} MB"
                     if tot_mb else f"Loading data from cloud… {mb:.0f} MB")
            _set_job(progress=done, total=total or done, stage=stage)

        _set_job(stage="Loading data from cloud…", progress=0, total=0)
        snap_bytes = _sb_download(_SB_SNAP, on_progress=dl_progress)
        if not snap_bytes:
            _set_job(status="idle", stage="", progress=0, total=0)
            with _SB_AUTOLOAD_LOCK:
                _SB_AUTOLOAD_DONE = False
            return

        _set_job(stage="Restoring dataset…", progress=0, total=0)
        ds = _unpack_ds(snap_bytes)
        STATE["ds"] = ds
        n = len(ds["records"])
        _set_job(status="done", progress=n, total=n, stage="")
    except Exception:
        _set_job(status="error", error=traceback.format_exc())


def _try_autoload():
    """Called from / when STATE[ds] is None and job is idle.
    Checks Supabase once per process lifetime; starts background load if a
    stored masterlist is found. Returns True if a background load was started."""
    global _SB_AUTOLOAD_DONE
    if not _sb_ok():
        return False
    with _SB_AUTOLOAD_LOCK:
        if _SB_AUTOLOAD_DONE:
            return False
        _SB_AUTOLOAD_DONE = True
    try:
        meta = _sb_get_meta()
    except Exception:
        return False
    if not meta:
        return False
    print("[supabase] stored masterlist found — auto-reloading", flush=True)
    _set_job(status="parsing", progress=0, total=0,
             stage="Auto-reloading from cloud storage…", error=None)
    threading.Thread(target=_do_load_from_supabase, args=(meta,), daemon=True).start()
    return True


# ── Inlined Chart.js (for the shell) ─────────────────────────────────────────
_SCRIPT_CACHE = {}
_CDN_SCRIPTS = [
    "https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js",
    "https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js",
]


def fetch_scripts():
    out = []
    for url in _CDN_SCRIPTS:
        if url not in _SCRIPT_CACHE:
            try:
                with urllib.request.urlopen(url, timeout=10) as r:
                    _SCRIPT_CACHE[url] = r.read().decode("utf-8")
            except Exception:
                _SCRIPT_CACHE[url] = None
        out.append(_SCRIPT_CACHE[url])
    return out


_LOGO = None


def load_logo():
    global _LOGO
    if _LOGO:
        return _LOGO
    base = os.path.dirname(os.path.abspath(__file__))
    for name, mime in (("logo.png", "image/png"), ("logo.jpg", "image/jpeg"),
                       ("logo.jpeg", "image/jpeg"), ("logo.webp", "image/webp")):
        p = os.path.join(base, name)
        try:
            if os.path.getsize(p) == 0:
                continue
            with open(p, "rb") as f:
                _LOGO = f"data:{mime};base64,{base64.b64encode(f.read()).decode('utf-8')}"
            return _LOGO
        except OSError:
            continue
    return ""


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_FULL = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# ── Business rules (ported from the original config.py) ──────────────────────
REVENUE_COMPONENTS = ["DUTY", "VAT", "EXCISEADVALOREM", "OTHERTAXES",
                      "DUTIESTAXES", "FINEPENALTIES", "GLOBALFEES", "TOTAL_NONCASH"]
OIL_INDEX2 = {"GASOLINE", "DIESEL", "JET FUEL", "LUBRICATING OIL",
              "OTHER OIL PRODUCTS", "LPG", "BUNKER FUEL"}
OIL_VALIDITY = {"OIL DV"}
VEHICLE_HS4 = {"8702", "8703", "8704"}
VEHICLE_INDEX2 = {"MOTOR VEHICLE EX", "MOTOR VEHICLE OTHERS", "TRUCKS"}
VEHICLE_TYPE = {"8702": "Buses/Vans 10+ pax (8702)", "8703": "Passenger Cars (8703)",
                "8704": "Trucks/Goods (8704)"}
CKD_RE = re.compile(r"(?:KD PARTS|CKD|SKD|COMPONENT|ASSEMBLY|SUB-?ASSY|\bASSY\b|KNOCK|SPARE PART|\bPARTS\b)")
_WS = re.compile(r"\s+")
_UNIT = re.compile(r"(\d[\d,]*)\s*\)?\s*UNIT")

# ── Vehicle classification by HS code ────────────────────────────────────────
# Passenger-car powertrain (chapter 8703, HS-6):
#   ICE      : 8703.21–8703.33 (conventional petrol + diesel cars)
#   Hybrid   : 8703.40–8703.70 (HEV / PHEV)
#   Electric : 8703.80 (BEV cars)
CAR_ICE_HS6 = {"870321", "870322", "870323", "870324", "870331", "870332", "870333"}
CAR_HYBRID_HS6 = {"870340", "870350", "870360", "870370"}
CAR_ELECTRIC_HS6 = {"870380"}

# Goods vehicles (chapter 8704): PICKUPS are grouped WITH cars and get a powertrain;
# everything else under 8704 is a TRUCK (own tab).
# NB: 8704.21/.31 (≤5 t) hold BOTH pickups (Hilux/Triton/Ranger) AND light
# cab-chassis trucks (Hino/Fuso/L300). They are split by the 8-DIGIT subheading:
# pickups are the "…26" / "…12" lines; the "…19" / "…29" lines are light trucks.
PICKUP_HS8 = {"87042112", "87042126",    # 8704.21 diesel pickups (Hilux/Triton/D-Max)
              "87043126",                # 8704.31 petrol pickup (Ranger)
              "87044126", "87045126",    # 8704.41/.51 hybrid pickups (Navara PHEV, etc.)
              "87046011", "87046021"}    # 8704.60 electric pickups
ELECTRIC_GOODS_HS6 = "870460"            # 8704.60 electric goods vehicles
PICKUP_HYBRID_HS6 = {"870441", "870451"}  # 8704.41/.51 hybrid pickups

# Brand-level view (Car Imports → 🏷️ By Brand sub-tab): wider than the strict
# Pickup/Truck split — covers ALL passenger cars (8703) plus EVERY ≤5t goods-vehicle
# subheading (8704.21/.31/.41/.51/.60), so light trucks/vans that the 8-digit split
# routes to the Trucks tab (e.g. the Toyota Lite Ace, HS 87043129) are included here.
LIGHT_GOODS_HS6 = {"870421", "870431", "870441", "870451", "870460"}

CAR_BRANDS = ["TOYOTA", "MITSUBISHI", "ISUZU", "HINO", "NISSAN", "HONDA", "SUZUKI", "MAZDA",
              "FORD", "HYUNDAI", "KIA", "CHEVROLET", "SUBARU", "VOLKSWAGEN", "PEUGEOT",
              "RENAULT", "BMW", "MERCEDES-BENZ", "AUDI", "LEXUS", "VOLVO", "JEEP", "CHANGAN",
              "GEELY", "GAC", "MG", "CHERY", "FOTON", "JAC", "DFSK", "GREAT WALL", "HAVAL",
              "TANK", "BYD", "VINFAST", "TESLA", "WULING", "BAIC", "MAXUS", "SSANGYONG",
              "DONGFENG", "JMC", "KMC"]
_BRAND_RE = re.compile(r"\b(" + "|".join(re.escape(b) for b in CAR_BRANDS) + r")\b")


def detect_brand(prod, importer):
    m = _BRAND_RE.search((prod or "").upper())
    if m:
        return m.group(1)
    m = _BRAND_RE.search((importer or "").upper())
    if m:
        return m.group(1)
    return "Other / unidentified"


def _light_pt(hs6):
    """Powertrain for the brand view — same HS-code rules as build_records(), just
    computed independently of the stored PT field so it also covers light goods
    vehicles (Lite Ace and the like) that build_records() leaves PT=None for."""
    if hs6 in CAR_ELECTRIC_HS6 or hs6 == ELECTRIC_GOODS_HS6:
        return "Electric"
    if hs6 in CAR_HYBRID_HS6 or hs6 in PICKUP_HYBRID_HS6:
        return "Hybrid"
    return "ICE"


# Public Transport vs Private/Other (Car Imports → 🚌 Motor Vehicles sub-tab):
# the masterlist already classifies this directly — GENERAL DESC "PUBLIC TRANSPORT"
# is a distinct bucket from "MOTOR VEHICLES" (confirmed against the real 2026 file:
# e.g. Toyota HiAce / Nissan Urvan vans, HS 8702.10.99, land under GENERAL DESC =
# SPECIFIC DESC = "PUBLIC TRANSPORT"). No heuristic needed — just read the column.
MOTOR_VEHICLE_GENS = {"MOTOR VEHICLES", "PUBLIC TRANSPORT"}

# Within GENERAL DESC "MOTOR VEHICLES", SPECIFIC DESC "PARTS & ACCESSORIES" is parts/
# components, not an actual vehicle — keep it out of the Motor Vehicles tab and give
# it its own sub-tab instead.
MOTOR_VEHICLE_PARTS_SPEC = "PARTS & ACCESSORIES"


def is_public_transport(gen):
    return gen == "PUBLIC TRANSPORT"


# Pseudo-fields: computed per-record rather than read from a stored tuple column.
# api_drill()/api_entries() special-case these (group-by AND constrain-by both work).
PSEUDO_FIELDS = {
    "brand": lambda r: detect_brand(r[PROD], r[IMP]),
    "pubtransport": lambda r: "Public Transport" if is_public_transport(r[GEN]) else "Private / Other",
}


# ── Record tuple schema ──────────────────────────────────────────────────────
(ENTRY, ORD, REV, VOL, PORT, IMP, I2, GEN, PROD, ORIG, HS4, HS6, HS11,
 ISOIL, ISVEH, ISCKD, VTYPE, PT, UNITS, QTY, COMPS, DSTR, VCLASS, HS8, SPEC) = range(25)

DATE_BASES = ("COLLECTIONDATE", "ASSESSMENTDATE", "REGISTRYDATE")

# Global in-memory dataset (single-user local tool). "job" tracks the background
# parse kicked off by /upload — parsing a 200k+ row masterlist can take a couple
# minutes, far longer than a host's reverse-proxy will hold a request open, so
# /upload returns almost instantly and the upload page polls /upload_status
# instead of blocking on the parse.
STATE = {"ds": None,
         "job": {"status": "idle", "progress": 0, "total": 0, "stage": "", "error": None}}
JOB_LOCK = threading.Lock()


def _set_job(**kw):
    with JOB_LOCK:
        STATE["job"].update(kw)


def _get_job():
    with JOB_LOCK:
        return dict(STATE["job"])


class _NamedBytes(io.BytesIO):
    """io.BytesIO with a .filename so it satisfies the FileStorage-shaped
    interface load_targets() expects, without holding a Flask request open."""
    def __init__(self, data, filename):
        super().__init__(data)
        self.filename = filename


# ── Coercion helpers ─────────────────────────────────────────────────────────

def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def clean_text(v):
    return "" if v is None else _WS.sub(" ", str(v)).strip()


def norm_name(v):
    return clean_text(v).upper()


def parse_iso(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, f).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s[:19]).strftime("%Y-%m-%d")
    except ValueError:
        return None


_intern = sys.intern


def build_records(file_storage, date_basis, on_progress=None):
    """Parse the uploaded xlsx directly into compact, interned tuples sorted by
    date ordinal — single pass over the sheet (the old version built a full
    220k-row list of dicts first, then made a second pass over that list;
    merging the two saves one full materialization of the row data, though
    `dict.get` per row turned out to be the cheap part — openpyxl's own XML
    parsing per row dominates the wall-clock time either way). If given,
    on_progress(i, total) is called periodically with the current/total row
    count so a caller (e.g. a background upload thread) can report status."""
    import openpyxl
    wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    try:
        total = ws.max_row - 1 if ws.max_row else 0
    except Exception:
        total = 0

    recs = []
    undated = 0
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else f"col_{j}" for j, h in enumerate(row)]
            continue
        if on_progress and i % 5000 == 0:
            on_progress(i, total)
        if row is None or all(v is None for v in row):
            continue
        g = dict(zip(headers, row)).get

        iso = parse_iso(g(date_basis))
        if not iso:
            undated += 1
            continue
        ordv = date(int(iso[:4]), int(iso[5:7]), int(iso[8:10])).toordinal()
        rev = safe_float(g("TOTALASSESSMENT"))
        vol = safe_float(g("GROSSMASS"))
        port = _intern(clean_text(g("PORT")) or "(no port)")
        importer = _intern(norm_name(g("CONSIGNEE")) or "(unknown)")
        i2 = _intern(clean_text(g("INDEX2")).upper() or "UNCLASSIFIED")
        gen = _intern(clean_text(g("GENERAL DESC")) or "(unclassified)")
        spec = _intern(clean_text(g("SPECIFIC DESC")) or "(unspecified)")
        prod = clean_text(g("GOODS_DESCRIPTION")) or "(no description)"
        orig = _intern(clean_text(g("COUNTRYORIGIN")) or "(n/a)")
        hsd = re.sub(r"\D", "", str(g("HSCODE") or ""))   # digits-only HS code
        hs4 = _intern(hsd[:4]); hs6 = _intern(hsd[:6]); hs8 = _intern(hsd[:8])
        hs11 = clean_text(g("HS11_DSC")) or "(unlabeled)"
        desc_u = norm_name(g("GOODS_DESCRIPTION"))
        is_oil = (norm_name(g("validity")) in OIL_VALIDITY) or (i2 in OIL_INDEX2)
        is_veh = (hs4 in VEHICLE_HS4) or (i2 in VEHICLE_INDEX2)
        vtype = _intern(VEHICLE_TYPE[hs4]) if (is_veh and hs4 in VEHICLE_TYPE) else None
        is_ckd = is_veh and bool(CKD_RE.search(desc_u))
        units = 0.0
        pt = None
        vclass = None
        if is_veh and not is_ckd:
            q = safe_float(g("QUANTITY"))
            units = q if q > 0 else 0.0
            if units <= 0:
                m = _UNIT.search(desc_u)
                if m:
                    units = safe_float(m.group(1))
            # goods-vehicle body class (chapter 8704): pickup (specific HS-8) vs truck
            if hs8 in PICKUP_HS8:
                vclass = "Pickup"
            elif hs4 == "8704":
                vclass = "Truck"          # all other 8704 goods vehicles = trucks
            # powertrain — cars (8703) + pickups; trucks get no powertrain (own tab)
            if vclass == "Truck":
                pt = None
            elif hs6 in CAR_ELECTRIC_HS6:
                pt = "Electric"
            elif hs6 in CAR_HYBRID_HS6:
                pt = "Hybrid"
            elif hs6 in CAR_ICE_HS6:
                pt = "ICE"
            elif vclass == "Pickup":
                if hs6 == ELECTRIC_GOODS_HS6:
                    pt = "Electric"
                elif hs6 in PICKUP_HYBRID_HS6:
                    pt = "Hybrid"
                else:
                    pt = "ICE"
            pt = _intern(pt) if pt else None
            vclass = _intern(vclass) if vclass else None
        comps = tuple(safe_float(g(c)) for c in REVENUE_COMPONENTS)
        recs.append((_intern(clean_text(g("ENTRY_CODE")) or "(none)"), ordv, rev, vol,
                     port, importer, i2, gen, prod, orig, hs4, hs6, hs11,
                     is_oil, is_veh, is_ckd, vtype, pt, units, safe_float(g("QUANTITY")),
                     comps, iso, vclass, hs8, spec))
    wb.close()
    recs.sort(key=lambda r: r[ORD])
    if on_progress:
        on_progress(total, total)
    return recs, undated


# ── Targets ──────────────────────────────────────────────────────────────────

def load_targets(file_storage=None):
    if file_storage is not None and getattr(file_storage, "filename", ""):
        text = file_storage.read().decode("utf-8-sig", errors="replace")
    else:
        p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "targets.csv")
        if not os.path.exists(p):
            return {}
        with open(p, "r", encoding="utf-8-sig") as f:
            text = f.read()
    out = {}
    for r in csv.DictReader(io.StringIO(text)):
        try:
            out[(int(float(r["year"])), int(float(r["month"])))] = \
                float(str(r["monthly_target_php"]).replace(",", "") or 0)
        except (TypeError, ValueError, KeyError):
            continue
    return out


def days_in_month(y, m):
    nxt = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
    return (nxt - date(y, m, 1)).days


def target_for_range(targets, start, end):
    if not targets or start > end:
        return 0.0
    total = 0.0
    cur = date(start.year, start.month, 1)
    while cur <= end:
        dim = days_in_month(cur.year, cur.month)
        lo, hi = max(start, cur), min(end, date(cur.year, cur.month, dim))
        if hi >= lo:
            total += targets.get((cur.year, cur.month), 0.0) * ((hi - lo).days + 1) / dim
        cur = date(cur.year + 1, 1, 1) if cur.month == 12 else date(cur.year, cur.month + 1, 1)
    return total


def add_months(d, n):
    m = d.month - 1 + n
    y, m = d.year + m // 12, m % 12 + 1
    return date(y, m, min(d.day, days_in_month(y, m)))


def pct_change(cur, base):
    return None if not base else round((cur - base) / base * 100.0, 1)


def fmt_pct(v):
    return "n.a." if v is None else f"{'+' if v >= 0 else ''}{v:.1f}%"


# ── Query engine ─────────────────────────────────────────────────────────────

class F:
    """A filter spec for one query."""
    __slots__ = ("oil", "ports", "domain", "pt", "cons")

    def __init__(self, oil=None, ports=None, domain=None, pt=None, cons=None):
        self.oil = oil if oil in ("Oil", "Non-oil") else None
        self.ports = set(ports) if ports else None
        self.domain = domain          # None | 'vehicle' | 'vehicle_ckd' | 'oil'
        self.pt = pt if pt in ("Electric", "Hybrid", "ICE") else None
        self.cons = cons or []        # list of (field_index, value)


def _slice(ds, start_ord, end_ord):
    lo = bisect.bisect_left(ds["ords"], start_ord)
    hi = bisect.bisect_right(ds["ords"], end_ord)
    return ds["records"][lo:hi]


def _iter(slc, f):
    oil, ports, dom, pt, cons = f.oil, f.ports, f.domain, f.pt, f.cons
    for r in slc:
        if oil == "Oil" and not r[ISOIL]:
            continue
        if oil == "Non-oil" and r[ISOIL]:
            continue
        if ports and r[PORT] not in ports:
            continue
        if dom == "vehicle":
            if not r[ISVEH] or r[ISCKD]:
                continue
        elif dom == "vehicle_ckd":
            if not r[ISVEH] or not r[ISCKD]:
                continue
        elif dom == "car":            # passenger cars + pickups (have a powertrain)
            if r[PT] is None:
                continue
        elif dom == "truck":          # medium/heavy goods vehicles
            if r[VCLASS] != "Truck":
                continue
        elif dom == "carbrand":       # brand view: cars + ALL ≤5t goods vehicles (incl. Lite Ace-likes)
            if not (r[HS4] == "8703" or r[HS6] in LIGHT_GOODS_HS6):
                continue
        elif dom == "motorveh":       # GENERAL DESC = MOTOR VEHICLES or PUBLIC TRANSPORT, excl. parts
            if r[GEN] not in MOTOR_VEHICLE_GENS or r[SPEC] == MOTOR_VEHICLE_PARTS_SPEC:
                continue
        elif dom == "motorvehparts":  # the parts/accessories carve-out from the above
            if r[GEN] not in MOTOR_VEHICLE_GENS or r[SPEC] != MOTOR_VEHICLE_PARTS_SPEC:
                continue
        elif dom == "oil" and not r[ISOIL]:
            continue
        if pt and r[PT] != pt:
            continue
        ok = True
        for idx, val in cons:
            if r[idx] != val:
                ok = False
                break
        if ok:
            yield r


def grp(slc, f, field):
    g = {}
    for r in _iter(slc, f):
        k = r[field]
        e = g.get(k)
        if e is None:
            g[k] = [r[REV], r[VOL], {r[ENTRY]}, 1, r[UNITS]]
        else:
            e[0] += r[REV]; e[1] += r[VOL]; e[2].add(r[ENTRY]); e[3] += 1; e[4] += r[UNITS]
    return g


def grp_list(slc, f, field, n=None, order="revenue", units=False):
    g = grp(slc, f, field)
    out = [{"name": k, "revenue": round(v[0], 2), "volume": round(v[1], 2),
            "entries": len(v[2]), "lines": v[3], "units": round(v[4], 0)}
           for k, v in g.items()]
    key = "units" if units else ("volume" if order == "volume" else "revenue")
    out.sort(key=lambda x: x[key], reverse=True)
    return out[:n] if n else out


def grp_list_keyfn(records, keyfn, n=None, order="revenue"):
    """Like grp_list, but groups by an arbitrary computed key (e.g. detect_brand())
    instead of a stored tuple field — records is a plain pre-filtered list."""
    g = {}
    for r in records:
        k = keyfn(r)
        e = g.get(k)
        if e is None:
            g[k] = [r[REV], r[VOL], {r[ENTRY]}, 1, r[UNITS]]
        else:
            e[0] += r[REV]; e[1] += r[VOL]; e[2].add(r[ENTRY]); e[3] += 1; e[4] += r[UNITS]
    out = [{"name": k, "revenue": round(v[0], 2), "volume": round(v[1], 2),
            "entries": len(v[2]), "lines": v[3], "units": round(v[4], 0)}
           for k, v in g.items()]
    key = "volume" if order == "volume" else "revenue"
    out.sort(key=lambda x: x[key], reverse=True)
    return out[:n] if n else out


def totals(slc, f):
    rev = vol = units = 0.0
    lines = 0
    ents = set()
    for r in _iter(slc, f):
        rev += r[REV]; vol += r[VOL]; ents.add(r[ENTRY]); lines += 1; units += r[UNITS]
    return {"revenue": round(rev, 2), "volume": round(vol, 2),
            "entries": len(ents), "lines": lines, "units": round(units, 0)}


# ── Request param helpers ────────────────────────────────────────────────────

def _ds_or_400():
    ds = STATE["ds"]
    if ds is None:
        return None
    return ds


def _range(ds):
    """Return (start_date, end_date, start_ord, end_ord) clamped to data bounds."""
    s = request.args.get("start") or ds["dmin"].isoformat()
    e = request.args.get("end") or ds["dmax"].isoformat()
    try:
        sd = date.fromisoformat(s)
    except ValueError:
        sd = ds["dmin"]
    try:
        ed = date.fromisoformat(e)
    except ValueError:
        ed = ds["dmax"]
    return sd, ed, sd.toordinal(), ed.toordinal()


def _filt(extra_domain=None, extra_pt=None, cons=None):
    oil = request.args.get("oil")
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    domain = extra_domain or request.args.get("domain")
    pt = extra_pt or request.args.get("pt")
    return F(oil=oil, ports=ports, domain=domain, pt=pt, cons=cons)


# ── API: summary / overview ──────────────────────────────────────────────────

@app.route("/api/summary")
def api_summary():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    slc = _slice(ds, so, eo)
    f = _filt()
    t = totals(slc, f)
    t["n_importers"] = len(grp(slc, f, IMP))
    t["n_commodities"] = len(grp(slc, f, I2))
    t["n_ports"] = len(grp(slc, f, PORT))
    t["avg_per_entry"] = round(t["revenue"] / t["entries"], 2) if t["entries"] else 0

    # monthly series (rev + prorated target) + daily series
    months = {}
    daily = {}
    for r in _iter(slc, f):
        ym = r[DSTR][:7]
        m = months.get(ym)
        if m is None:
            months[ym] = [r[REV], r[VOL]]
        else:
            m[0] += r[REV]; m[1] += r[VOL]
        dd = daily.get(r[DSTR])
        if dd is None:
            daily[r[DSTR]] = [r[REV], r[VOL]]
        else:
            dd[0] += r[REV]; dd[1] += r[VOL]
    monthly = []
    for ym in sorted(months):
        y, mo = int(ym[:4]), int(ym[5:7])
        dim = days_in_month(y, mo)
        ov_lo, ov_hi = max(sd, date(y, mo, 1)), min(ed, date(y, mo, dim))
        tgt = target_for_range(ds["targets"], ov_lo, ov_hi)
        monthly.append({"ym": ym, "label": f"{MONTHS[mo - 1]} {y}",
                        "revenue": round(months[ym][0], 2), "target": round(tgt, 2)})
    dseries = [{"date": d, "revenue": round(daily[d][0], 2)} for d in sorted(daily)]

    fo = F(oil=None, ports=f.ports)
    oil_pie = {"Oil": 0.0, "Non-oil": 0.0}
    for r in _iter(slc, fo):
        oil_pie["Oil" if r[ISOIL] else "Non-oil"] += r[REV]
    oil_pie = {k: round(v, 2) for k, v in oil_pie.items()}

    return jsonify(
        totals=t, monthly=monthly, daily=dseries, oil_pie=oil_pie,
        top_importers=grp_list(slc, f, IMP, 10),
        top_commodities=grp_list(slc, f, I2, 10),
        span={"start": sd.isoformat(), "end": ed.isoformat()},
    )


@app.route("/api/collection")
def api_collection():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    f = _filt()
    slc = _slice(ds, so, eo)
    pt = totals(slc, f)
    tgt = target_for_range(ds["targets"], sd, ed)
    period = {"start": sd.isoformat(), "end": ed.isoformat(), "revenue": pt["revenue"],
              "volume": pt["volume"], "entries": pt["entries"], "target": round(tgt, 2),
              "variance": round(pt["revenue"] - tgt, 2), "variance_pct": pct_change(pt["revenue"], tgt)}

    yst = date(ed.year, 1, 1)
    yslc = _slice(ds, yst.toordinal(), eo)
    yrev = totals(yslc, f)["revenue"]
    ytgt = target_for_range(ds["targets"], yst, ed)
    ytd = {"start": yst.isoformat(), "revenue": yrev, "target": round(ytgt, 2),
           "variance": round(yrev - ytgt, 2), "variance_pct": pct_change(yrev, ytgt)}

    # cumulative daily over range
    daily = {}
    comps = [0.0] * len(REVENUE_COMPONENTS)
    for r in _iter(slc, f):
        daily[r[DSTR]] = daily.get(r[DSTR], 0.0) + r[REV]
        rc = r[COMPS]
        for i in range(len(comps)):
            comps[i] += rc[i]
    cum = 0.0
    cumlist = []
    for d in sorted(daily):
        cum += daily[d]
        cumlist.append({"date": d, "cum": round(cum, 2)})
    components = [{"component": REVENUE_COMPONENTS[i], "amount": round(comps[i], 2)}
                 for i in range(len(comps)) if comps[i]]

    yoy = None
    if ds.get("prior"):
        ps, pe = add_months(sd, -12), add_months(ed, -12)
        pslc = _slice(ds["prior"], ps.toordinal(), pe.toordinal())
        prev = totals(pslc, f)["revenue"]
        yoy = {"prior_rev": prev, "rev": pct_change(pt["revenue"], prev),
               "efficiency": round(pt["revenue"] / prev * 100, 1) if prev else None,
               "label": f"{ps.isoformat()} → {pe.isoformat()}"}

    return jsonify(period=period, ytd=ytd, cumulative=cumlist, components=components, yoy=yoy)


@app.route("/api/oil")
def api_oil():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]

    def split(slc):
        d = {"Oil": [0.0, 0.0], "Non-oil": [0.0, 0.0]}
        for r in _iter(slc, F(oil=None, ports=ports)):
            k = "Oil" if r[ISOIL] else "Non-oil"
            d[k][0] += r[REV]; d[k][1] += r[VOL]
        return d

    cur = split(_slice(ds, so, eo))
    ms, me = add_months(sd, -1), add_months(ed, -1)
    mom = split(_slice(ds, ms.toordinal(), me.toordinal()))
    yoy = None
    if ds.get("prior"):
        ys, ye = add_months(sd, -12), add_months(ed, -12)
        yoy = split(_slice(ds["prior"], ys.toordinal(), ye.toordinal()))

    out = {}
    for cls in ("Oil", "Non-oil"):
        c = cur[cls]
        out[cls] = {"revenue": round(c[0], 2), "volume": round(c[1], 2),
                    "rev_mom": pct_change(c[0], mom[cls][0]), "vol_mom": pct_change(c[1], mom[cls][1]),
                    "rev_yoy": pct_change(c[0], yoy[cls][0]) if yoy else None,
                    "vol_yoy": pct_change(c[1], yoy[cls][1]) if yoy else None}
    return jsonify(oil=out, has_prior=bool(ds.get("prior")),
                   mom_label=f"{ms.isoformat()} → {me.isoformat()}")


def _trends(ds, field, n=50):
    sd, ed, so, eo = _range(ds)
    order = request.args.get("order", "revenue")
    f = _filt()
    cur = grp_list(_slice(ds, so, eo), f, field, n, order)
    names = {r["name"] for r in cur}
    ms, me = add_months(sd, -1), add_months(ed, -1)
    momg = grp(_slice(ds, ms.toordinal(), me.toordinal()), f, field)
    yoyg = None
    if ds.get("prior"):
        ys, ye = add_months(sd, -12), add_months(ed, -12)
        yoyg = grp(_slice(ds["prior"], ys.toordinal(), ye.toordinal()), f, field)
    for r in cur:
        nm = r["name"]
        pm = momg.get(nm)
        r["rev_mom"] = pct_change(r["revenue"], pm[0] if pm else 0)
        r["vol_mom"] = pct_change(r["volume"], pm[1] if pm else 0)
        r["new"] = (pm is None)
        if yoyg is not None:
            yg = yoyg.get(nm)
            r["rev_yoy"] = pct_change(r["revenue"], yg[0]) if yg else None
            r["vol_yoy"] = pct_change(r["volume"], yg[1]) if yg else None
        else:
            r["rev_yoy"] = r["vol_yoy"] = None
    return cur


@app.route("/api/importers")
def api_importers():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    return jsonify(rows=_trends(ds, IMP, 60), has_prior=bool(ds.get("prior")))


@app.route("/api/commodities")
def api_commodities():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    field = GEN if request.args.get("field") == "gd" else I2
    return jsonify(rows=_trends(ds, field, 60), has_prior=bool(ds.get("prior")))


# ── Generic drill-down + entries ─────────────────────────────────────────────

_FIELD = {"importer": IMP, "commodity": I2, "general": GEN, "specific": SPEC, "product": PROD,
          "port": PORT, "origin": ORIG, "hs6": HS6, "hs11": HS11, "hs4": HS4, "hs8": HS8}


@app.route("/api/drill")
def api_drill():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    cf_name = request.args.get("child_field")
    order = request.args.get("order", "revenue")
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    oil = request.args.get("oil")
    dom = request.args.get("domain")
    ptf = request.args.get("pt")
    cons = []
    # legacy single parent (parent_field/parent_value); oilclass = an oil filter
    pfname = request.args.get("parent_field")
    pv = request.args.get("parent_value", "")
    if pfname == "oilclass":
        oil = pv
    elif pfname in _FIELD:
        cons.append((_FIELD[pfname], pv))
    # chained named constraints (deep drill: general/specific/product/importer/…)
    for key in ("general", "specific", "product", "importer", "commodity",
                "origin", "hs4", "hs6", "hs8", "hs11", "port"):
        v = request.args.get(key)
        if v is not None:
            cons.append((_FIELD[key], v))
    # pseudo-field constraints (brand/pubtransport): not stored tuple columns, computed on the fly
    pseudo_cons = {k: request.args.get(k) for k in PSEUDO_FIELDS if request.args.get(k) is not None}
    f = F(oil=oil, ports=ports, domain=dom, pt=ptf, cons=cons)
    slc = _slice(ds, so, eo)

    if cf_name in PSEUDO_FIELDS or pseudo_cons:
        rows = list(_iter(slc, f))
        for key, val in pseudo_cons.items():
            keyfn = PSEUDO_FIELDS[key]
            rows = [r for r in rows if keyfn(r) == val]
        if cf_name in PSEUDO_FIELDS:
            return jsonify(rows=grp_list_keyfn(rows, PSEUDO_FIELDS[cf_name], 80, order))
        cf = _FIELD.get(cf_name)
        if cf is None:
            return jsonify(error="bad fields"), 400
        return jsonify(rows=grp_list(rows, F(), cf, 80, order))

    cf = _FIELD.get(cf_name)
    if cf is None:
        return jsonify(error="bad fields"), 400
    return jsonify(rows=grp_list(slc, f, cf, 80, order))


@app.route("/api/entries")
def api_entries():
    """Distinct entry (declaration) numbers for the constrained selection."""
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    cons = []
    for key in ("importer", "commodity", "general", "specific", "product", "port", "origin",
                "hs6", "hs11", "hs4", "hs8"):
        v = request.args.get(key)
        if v is not None:
            cons.append((_FIELD[key], v))
    pseudo_cons = {k: request.args.get(k) for k in PSEUDO_FIELDS if request.args.get(k) is not None}
    f = _filt(cons=cons)
    agg = {}
    for r in _iter(_slice(ds, so, eo), f):
        if any(PSEUDO_FIELDS[k](r) != v for k, v in pseudo_cons.items()):
            continue
        e = agg.get(r[ENTRY])
        if e is None:
            agg[r[ENTRY]] = [r[REV], r[VOL], 1, r[DSTR], r[PORT], r[IMP]]
        else:
            e[0] += r[REV]; e[1] += r[VOL]; e[2] += 1
    out = [{"entry": k, "revenue": round(v[0], 2), "volume": round(v[1], 2),
            "lines": v[2], "date": v[3], "port": v[4], "importer": v[5]}
           for k, v in agg.items()]
    out.sort(key=lambda x: x["revenue"], reverse=True)
    return jsonify(rows=out[:1000], total=len(out))


@app.route("/api/export_raw")
def api_export_raw():
    """Raw line items for the current period/oil/port filter, string-interned for
    size, so the static-report export can run the drill-down + entries logic
    fully offline (see decodeExport()/localApi() in dashboard.html). Field order
    in each row: ENTRY,REV,VOL,PORT,IMP,I2,GEN,SPEC,PROD,ORIG,HS4,HS6,HS8,HS11,
    ISOIL,ISVEH,ISCKD,VCLASS,PT,UNITS,DSTR (VCLASS/PT use -1 for null)."""
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    f = _filt()
    strtab, sidx = [], {}

    def s(v):
        v = v or ""
        i = sidx.get(v)
        if i is None:
            i = len(strtab)
            strtab.append(v)
            sidx[v] = i
        return i

    rows = []
    for r in _iter(_slice(ds, so, eo), f):
        rows.append([
            s(r[ENTRY]), r[REV], r[VOL], s(r[PORT]), s(r[IMP]), s(r[I2]), s(r[GEN]), s(r[SPEC]),
            s(r[PROD]), s(r[ORIG]), s(r[HS4]), s(r[HS6]), s(r[HS8]), s(r[HS11]),
            r[ISOIL], r[ISVEH], r[ISCKD], s(r[VCLASS]) if r[VCLASS] else -1,
            s(r[PT]) if r[PT] else -1, r[UNITS], s(r[DSTR]),
        ])
    return jsonify(strs=strtab, rows=rows)


# ── Cars ─────────────────────────────────────────────────────────────────────

@app.route("/api/cars")
def api_cars():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    ptf = request.args.get("powertrain")
    slc = _slice(ds, so, eo)

    # mix = powertrain split of passenger cars + pickups (trucks/buses excluded —
    # trucks live in the Trucks sub-tab).
    mixg = grp(slc, F(ports=ports, domain="car"), PT)
    mix = [{"name": k, "units": round(v[4], 0), "entries": len(v[2]), "duties": round(v[0], 2)}
           for k, v in mixg.items() if k]
    order = {"Electric": 0, "Hybrid": 1, "ICE": 2}
    mix.sort(key=lambda x: order.get(x["name"], 9))
    total_units = sum(round(v[4], 0) for v in mixg.values())   # cars + pickups

    fv = F(ports=ports, domain="car", pt=ptf)
    by_port_g = grp(slc, fv, PORT)
    by_port = [{"port": k, "units": round(v[4], 0), "entries": len(v[2]),
                "duties": round(v[0], 2), "gross_kg": round(v[1], 0)}
               for k, v in by_port_g.items()]
    by_port.sort(key=lambda x: x["units"], reverse=True)

    # powertrain by port (units)
    pbp = {}
    for r in _iter(slc, F(ports=ports, domain="car")):
        d = pbp.setdefault(r[PORT], {"Electric": 0.0, "Hybrid": 0.0, "ICE": 0.0})
        if r[PT]:
            d[r[PT]] += r[UNITS]
    powertrain_by_port = [{"port": p, **{k: round(v, 0) for k, v in d.items()}}
                          for p, d in sorted(pbp.items(), key=lambda x: -sum(x[1].values()))]

    top = grp_list(slc, fv, IMP, 40, units=True)
    # CKD / knock-down parts (excluded from unit counts): show declarations + line items
    ckd_g = grp(slc, F(ports=ports, domain="vehicle_ckd"), IMP)
    ckd = [{"name": k, "entries": len(v[2]), "lines": v[3], "duties": round(v[0], 2)}
           for k, v in sorted(ckd_g.items(), key=lambda x: -x[1][3])[:30]]

    return jsonify(mix=mix, total_units=round(total_units, 0), by_port=by_port,
                   powertrain_by_port=powertrain_by_port, top_importers=top, ckd=ckd,
                   powertrain=ptf or "All")


@app.route("/api/car_products")
def api_car_products():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    ptf = request.args.get("powertrain")
    imp = request.args.get("importer", "")
    f = F(ports=ports, domain="car", pt=ptf, cons=[(IMP, imp)])
    agg = {}
    for r in _iter(_slice(ds, so, eo), f):
        k = (r[PROD], r[PT] or "", r[VCLASS] or "Car", r[HS8], r[ORIG])
        e = agg.get(k)
        if e is None:
            agg[k] = [r[UNITS], 1, r[REV], {r[ENTRY]}]
        else:
            e[0] += r[UNITS]; e[1] += 1; e[2] += r[REV]; e[3].add(r[ENTRY])
    rows = [{"product": k[0], "powertrain": k[1], "vclass": k[2], "hs8": k[3], "origin": k[4],
             "units": round(v[0], 0), "entries": len(v[3]), "duties": round(v[2], 2)}
            for k, v in agg.items()]
    rows.sort(key=lambda x: x["units"], reverse=True)
    return jsonify(rows=rows[:80])


# ── Trucks sub-tab (medium/heavy goods vehicles) ─────────────────────────────

@app.route("/api/trucks")
def api_trucks():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    slc = _slice(ds, so, eo)
    f = F(ports=ports, domain="truck")
    t = totals(slc, f)
    elec = sum(r[UNITS] for r in _iter(slc, F(ports=ports, domain="truck")) if r[PT] == "Electric"
               or r[HS6] == ELECTRIC_GOODS_HS6)
    totals_out = {"units": t["units"], "entries": t["entries"], "duties": t["revenue"],
                  "electric_units": round(elec, 0)}
    bp = grp(slc, f, PORT)
    by_port = [{"port": k, "units": round(v[4], 0), "entries": len(v[2]), "duties": round(v[0], 2)}
               for k, v in sorted(bp.items(), key=lambda x: -x[1][4])]
    top = [{"name": x["name"], "units": x["units"], "entries": x["entries"], "duties": x["revenue"]}
           for x in grp_list(slc, f, IMP, 40, units=True)]
    return jsonify(totals=totals_out, by_port=by_port, top_importers=top)


@app.route("/api/truck_products")
def api_truck_products():
    """Truck models for one importer (drill-down)."""
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    imp = request.args.get("importer", "")
    agg = {}
    for r in _iter(_slice(ds, so, eo), F(ports=ports, domain="truck", cons=[(IMP, imp)])):
        k = (r[PROD], r[HS8], r[ORIG])
        e = agg.get(k)
        if e is None:
            agg[k] = [r[UNITS], 1, r[REV], {r[ENTRY]}]
        else:
            e[0] += r[UNITS]; e[1] += 1; e[2] += r[REV]; e[3].add(r[ENTRY])
    rows = [{"product": k[0], "hs8": k[1], "origin": k[2],
             "units": round(v[0], 0), "entries": len(v[3]), "duties": round(v[2], 2)}
            for k, v in agg.items()]
    rows.sort(key=lambda x: x["units"], reverse=True)
    return jsonify(rows=rows[:80])


# ── Car brands (🏷️ By Brand sub-tab) ─────────────────────────────────────────

@app.route("/api/car_brands")
def api_car_brands():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    f = F(ports=ports, domain="carbrand")
    brands = {}
    for r in _iter(_slice(ds, so, eo), f):
        b = detect_brand(r[PROD], r[IMP])
        pt = _light_pt(r[HS6])
        e = brands.get(b)
        if e is None:
            e = {"Electric": 0.0, "Hybrid": 0.0, "ICE": 0.0, "revenue": 0.0, "entries": set()}
            brands[b] = e
        e[pt] += r[UNITS]
        e["revenue"] += r[REV]
        e["entries"].add(r[ENTRY])
    rows = []
    for b, e in brands.items():
        units = e["Electric"] + e["Hybrid"] + e["ICE"]
        rows.append({"brand": b, "units": round(units, 0), "electric": round(e["Electric"], 0),
                      "hybrid": round(e["Hybrid"], 0), "ice": round(e["ICE"], 0),
                      "revenue": round(e["revenue"], 2), "entries": len(e["entries"])})
    rows.sort(key=lambda x: -x["units"])
    total_units = sum(r["units"] for r in rows)
    mix = {k: round(sum(r[k.lower()] for r in rows), 0) for k in ("Electric", "Hybrid", "ICE")}
    return jsonify(rows=rows[:40], total_units=round(total_units, 0), mix=mix)


# ── Motor Vehicles general-desc view (🚌 sub-tab: public transport vs private) ──

@app.route("/api/motor_vehicles")
def api_motor_vehicles():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    f = F(ports=ports, domain="motorveh")
    slc = _slice(ds, so, eo)
    t = totals(slc, f)
    pt = {"Public Transport": [0.0, 0.0, set(), 0.0], "Private / Other": [0.0, 0.0, set(), 0.0]}
    for r in _iter(slc, f):
        k = PSEUDO_FIELDS["pubtransport"](r)
        e = pt[k]
        e[0] += r[REV]; e[1] += r[VOL]; e[2].add(r[ENTRY]); e[3] += r[UNITS]
    pubtransport = [{"class": k, "revenue": round(v[0], 2), "volume": round(v[1], 2),
                      "entries": len(v[2]), "units": round(v[3], 0)} for k, v in pt.items()]
    top_importers = grp_list(slc, f, IMP, 40)
    return jsonify(totals=t, pubtransport=pubtransport, top_importers=top_importers)


# ── Motor Vehicle Parts & Accessories (carve-out, not a motor vehicle itself) ──

@app.route("/api/motor_vehicle_parts")
def api_motor_vehicle_parts():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    f = F(ports=ports, domain="motorvehparts")
    slc = _slice(ds, so, eo)
    t = totals(slc, f)
    top_importers = grp_list(slc, f, IMP, 40)
    top_goods = grp_list(slc, f, PROD, 40)
    return jsonify(totals=t, top_importers=top_importers, top_goods=top_goods)


# ── Commodity Sectors (INDEX2 full breakdown) ─────────────────────────────────

@app.route("/api/sectors")
def api_sectors():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    f = _filt()
    slc = _slice(ds, so, eo)
    t = totals(slc, f)
    sectors = grp_list(slc, f, I2, None)
    total_rev = t["revenue"] or 1
    for s in sectors:
        s["share"] = round(s["revenue"] / total_rev * 100, 1)
    return jsonify(totals=t, sectors=sectors)


# ── Petroleum ────────────────────────────────────────────────────────────────

@app.route("/api/petroleum")
def api_petroleum():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    slc = _slice(ds, so, eo)
    f = F(ports=ports, domain="oil")
    t = totals(slc, f)
    totals_out = {"kg": t["volume"], "mt": round(t["volume"] / 1e3, 0),
                  "duties": t["revenue"], "shipments": t["entries"]}
    # categorize by SPECIFIC DESC (the actual fuel type) instead of the messy
    # INDEX2 "OTHER OIL PRODUCTS" catch-all bucket.
    cat = grp_list(slc, f, SPEC, None)
    by_category = [{"category": c["name"], "mt": round(c["volume"] / 1e3, 0),
                    "kg": c["volume"], "duties": c["revenue"], "shipments": c["entries"]}
                   for c in cat]
    portg = grp_list(slc, f, PORT, None)
    by_port = [{"port": c["name"], "mt": round(c["volume"] / 1e3, 0),
                "duties": c["revenue"], "shipments": c["entries"]} for c in portg]
    impg = grp_list(slc, f, IMP, 30)
    top = [{"importer": c["name"], "mt": round(c["volume"] / 1e3, 1), "kg": c["volume"],
            "duties": c["revenue"], "shipments": c["entries"]} for c in impg]
    return jsonify(totals=totals_out, by_category=by_category, by_port=by_port, top_importers=top)


@app.route("/api/petro_products")
def api_petro_products():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    sd, ed, so, eo = _range(ds)
    ports = [p for p in request.args.get("ports", "").split(",") if p]
    cat = request.args.get("category")
    imp = request.args.get("importer")
    cons = []
    if cat is not None:
        cons.append((SPEC, cat))      # category = SPECIFIC DESC (fuel type)
    if imp is not None:
        cons.append((IMP, imp))
    f = F(ports=ports, domain="oil", cons=cons)
    agg = {}
    for r in _iter(_slice(ds, so, eo), f):
        k = (r[SPEC], r[HS11], r[HS8], r[ORIG]) if imp is not None else (r[HS11], r[HS8])
        e = agg.get(k)
        if e is None:
            agg[k] = [r[VOL], r[REV], {r[ENTRY]}]
        else:
            e[0] += r[VOL]; e[1] += r[REV]; e[2].add(r[ENTRY])
    if imp is not None:
        rows = [{"category": k[0], "product": k[1], "hs8": k[2], "origin": k[3],
                 "mt": round(v[0] / 1e3, 1), "duties": round(v[1], 2), "shipments": len(v[2])}
                for k, v in agg.items()]
    else:
        rows = [{"product": k[0], "hs8": k[1], "mt": round(v[0] / 1e3, 1),
                 "duties": round(v[1], 2), "shipments": len(v[2])} for k, v in agg.items()]
    rows.sort(key=lambda x: x["mt"], reverse=True)
    return jsonify(rows=rows[:60])


# ── Memo (formal A-F memorandum, computed for the selected period) ────────────

def _memo_agg(slc):
    """ONE pass over a slice: total rev/vol, oil split, and per-importer /
    per-commodity rev+vol. No distinct-entry sets (the memo doesn't need them),
    which keeps it fast even over the full span."""
    rev = vol = 0.0
    oil = {"Oil": [0.0, 0.0], "Non-oil": [0.0, 0.0]}
    imp, com = {}, {}
    for r in slc:
        rv, vl = r[REV], r[VOL]
        rev += rv; vol += vl
        ok = oil["Oil"] if r[ISOIL] else oil["Non-oil"]
        ok[0] += rv; ok[1] += vl
        nm = r[IMP]
        e = imp.get(nm)
        if e is None:
            imp[nm] = [rv, vl]
        else:
            e[0] += rv; e[1] += vl
        c = r[I2]
        e2 = com.get(c)
        if e2 is None:
            com[c] = [rv, vl]
        else:
            e2[0] += rv; e2[1] += vl
    return {"rev": rev, "vol": vol, "oil": oil, "imp": imp, "com": com}


def _rev_sum(slc):
    return sum(r[REV] for r in slc)


def _top_rows(dct, n):
    rows = [{"name": k, "revenue": round(v[0], 2), "volume": round(v[1], 2)}
            for k, v in dct.items()]
    rows.sort(key=lambda x: -x["revenue"])
    return rows[:n]


def _month_label(d):
    return f"{MONTHS_FULL[d.month - 1]} {d.year}"


def _range_label(s, e):
    if s.year == e.year and s.month == e.month:
        return f"{MONTHS_FULL[s.month - 1]} {s.day}-{e.day}, {e.year}"
    return f"{MONTHS_FULL[s.month - 1]} {s.day}, {s.year} - {MONTHS_FULL[e.month - 1]} {e.day}, {e.year}"


def _vol_block(cur_oil, prev_oil):
    cur_no, cur_o = cur_oil["Non-oil"], cur_oil["Oil"]
    prev_no, prev_o = prev_oil["Non-oil"], prev_oil["Oil"]
    cur_vol = cur_no[1] + cur_o[1]
    prev_vol = prev_no[1] + prev_o[1]
    cur_rev = cur_no[0] + cur_o[0]
    prev_rev = prev_no[0] + prev_o[0]
    return {
        "has": True, "cur_vol": cur_vol, "prev_vol": prev_vol,
        "vol_pct": pct_change(cur_vol, prev_vol),
        "nonoil_vol_pct": pct_change(cur_no[1], prev_no[1]),
        "oil_vol_pct": pct_change(cur_o[1], prev_o[1]),
        "nonoil_rev_d": cur_no[0] - prev_no[0], "nonoil_rev_pct": pct_change(cur_no[0], prev_no[0]),
        "oil_rev_d": cur_o[0] - prev_o[0], "oil_rev_pct": pct_change(cur_o[0], prev_o[0]),
        "net_rev_d": cur_rev - prev_rev, "net_rev_pct": pct_change(cur_rev, prev_rev),
    }


def _attach_deltas(rows, momg, yoyg):
    for r in rows:
        pm = momg.get(r["name"])
        py = yoyg.get(r["name"]) if yoyg is not None else None
        r["rev_mom"] = pct_change(r["revenue"], pm[0] if pm else 0)
        r["vol_mom"] = pct_change(r["volume"], pm[1] if pm else 0)
        r["rev_yoy"] = pct_change(r["revenue"], py[0]) if py else None
        r["vol_yoy"] = pct_change(r["volume"], py[1]) if py else None


def _new_growers(cur_rows, prevg):
    new, grow = [], []
    for r in cur_rows:
        pv = prevg.get(r["name"])
        if pv is None or pv[0] == 0:
            new.append({"name": r["name"], "rev": r["revenue"], "vol": r["volume"]})
        else:
            grow.append({"name": r["name"], "rev_d": r["revenue"] - pv[0],
                         "rev_pct": pct_change(r["revenue"], pv[0]),
                         "vol_d": r["volume"] - pv[1], "vol_pct": pct_change(r["volume"], pv[1])})
    new.sort(key=lambda x: -x["rev"])
    grow.sort(key=lambda x: -x["rev_d"])
    return new, grow


def _top_increases(cur_rows, prevg):
    out = []
    for r in cur_rows:
        pv = prevg.get(r["name"])
        pr, pvv = (pv[0], pv[1]) if pv else (0.0, 0.0)
        out.append({"name": r["name"], "rev_d": r["revenue"] - pr, "rev_pct": pct_change(r["revenue"], pr),
                    "vol_d": r["volume"] - pvv, "vol_pct": pct_change(r["volume"], pvv)})
    out.sort(key=lambda x: -x["rev_d"])
    return out


def compute_memo_ctx(ds):
    sd, ed, so, eo = _range(ds)
    prior = ds.get("prior")

    lm_s, lm_e = add_months(sd, -1), add_months(ed, -1)
    ly_s, ly_e = add_months(sd, -12), add_months(ed, -12)

    cur = _memo_agg(_slice(ds, so, eo))                         # 1 pass over the period
    lm = _memo_agg(_slice(ds, lm_s.toordinal(), lm_e.toordinal()))
    ly = _memo_agg(_slice(prior, ly_s.toordinal(), ly_e.toordinal())) if prior else None

    pt_rev = cur["rev"]
    tgt = target_for_range(ds["targets"], sd, ed)
    yst = date(ed.year, 1, 1)
    ytd_rev = _rev_sum(_slice(ds, yst.toordinal(), eo))
    ytd_tgt = target_for_range(ds["targets"], yst, ed)

    C = {"has_prior": bool(prior)}
    if prior:
        ply_rev = _rev_sum(_slice(prior, date(ed.year - 1, 1, 1).toordinal(),
                                  add_months(ed, -12).toordinal()))
        C.update(cur=ytd_rev, prev=ply_rev, diff=ytd_rev - ply_rev,
                 pct=pct_change(ytd_rev, ply_rev), eff=(ytd_rev / ply_rev * 100) if ply_rev else None)

    # daily revenue contributors — last 2 days of the period
    c_start = max(sd, ed - timedelta(days=1))
    contrib_rows = _top_rows(_memo_agg(_slice(ds, c_start.toordinal(), eo))["imp"], 40)
    contrib = {"label": _range_label(c_start, ed),
               "rows": [{"name": r["name"], "rev": r["revenue"]} for r in contrib_rows],
               "total": sum(r["revenue"] for r in contrib_rows)}

    D1 = _vol_block(cur["oil"], lm["oil"])
    D2 = {"has_prior": bool(prior)}
    if prior:
        D2.update(_vol_block(cur["oil"], ly["oil"]))

    # importers
    imp30 = _top_rows(cur["imp"], 30)
    ly_impg = ly["imp"] if prior else None
    _attach_deltas(imp30, lm["imp"], ly_impg)
    mom_new, mom_grow = _new_growers(imp30, lm["imp"])
    yoy_new, yoy_grow = _new_growers(imp30, ly_impg) if prior else ([], [])
    E = {"top30": imp30, "mom_new": mom_new, "mom_grow": mom_grow,
         "yoy_new": yoy_new, "yoy_grow": yoy_grow, "has_prior": bool(prior)}

    # commodities
    com30 = _top_rows(cur["com"], 30)
    ly_comg = ly["com"] if prior else None
    _attach_deltas(com30, lm["com"], ly_comg)
    Fc = {"top30": com30, "mom_top": _top_increases(com30, lm["com"]),
          "yoy_top": _top_increases(com30, ly_comg) if prior else [],
          "drivers": [r["name"] for r in imp30[:3]], "has_prior": bool(prior)}

    return {
        "year": ed.year,
        "period_label": _range_label(sd, ed),
        "end_label": f"{MONTHS_FULL[ed.month - 1]} {ed.day}, {ed.year}",
        "end_short": f"{MONTHS[ed.month - 1]} {ed.day}",
        "date_str": datetime.now().strftime("%d %B %Y"),
        "cur_month": _month_label(ed), "prev_month": _month_label(lm_e),
        "prev_year_month": _month_label(ly_e),
        "A": {"actual": pt_rev, "target": tgt, "diff": pt_rev - tgt,
              "dev_pct": pct_change(pt_rev, tgt)},
        "B": {"actual": ytd_rev, "target": ytd_tgt, "diff": ytd_rev - ytd_tgt,
              "dev_pct": pct_change(ytd_rev, ytd_tgt)},
        "C": C, "contrib": contrib, "D1": D1, "D2": D2, "E": E, "F": Fc,
    }


@app.route("/api/memo")
def api_memo():
    ds = _ds_or_400()
    if not ds:
        return jsonify(error="no data"), 400
    return jsonify(html=memo_builder.narrative_html(compute_memo_ctx(ds)))


@app.route("/api/memo_docx")
def api_memo_docx():
    ds = _ds_or_400()
    if not ds:
        return "No data", 400
    sd, ed, _, _ = _range(ds)
    data = memo_builder.build_docx(compute_memo_ctx(ds))
    resp = make_response(data)
    resp.headers["Content-Type"] = ("application/vnd.openxmlformats-officedocument."
                                    "wordprocessingml.document")
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="POB_Memo_{sd.isoformat()}_{ed.isoformat()}.docx"')
    return resp


# ── Routes: upload + shell ───────────────────────────────────────────────────

def _render_dashboard():
    ds = STATE["ds"]
    meta = {"file1": ds["file1"], "prior_name": ds["prior_name"],
            "date_basis": ds["date_basis"], "dmin": ds["dmin"].isoformat(), "dmax": ds["dmax"].isoformat(),
            "ports": ds["ports"], "buckets": ds["buckets"], "undated": ds["undated"],
            "has_prior": bool(ds["prior"]), "line_items": len(ds["records"]),
            "generated": datetime.now().strftime("%B %d, %Y at %H:%M")}
    chartjs, datalabels = fetch_scripts()
    html = render_template("dashboard.html", meta=json.dumps(meta),
                           chartjs_inline=chartjs, datalabels_inline=datalabels,
                           logo_url=load_logo())
    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp


@app.route("/")
def index():
    if STATE["ds"] is not None:
        return _render_dashboard()
    job = _get_job()
    if job["status"] == "parsing":
        return render_template("uploading.html")
    # First visit after a cold start: check Supabase for a stored masterlist
    if _try_autoload():
        return render_template("uploading.html")
    return render_template("index.html")


@app.route("/replace")
def replace():
    """Clear in-memory data and reset the auto-load flag so the user can
    upload a new masterlist. The next upload will overwrite Supabase too."""
    global _SB_AUTOLOAD_DONE
    STATE["ds"] = None
    _set_job(status="idle", progress=0, total=0, stage="", error=None)
    with _SB_AUTOLOAD_LOCK:
        _SB_AUTOLOAD_DONE = False
    from flask import redirect
    return redirect("/")


def _run_upload_job(f1_bytes, f1_name, f2_bytes, f2_name, targets_bytes, date_basis,
                    _save_to_sb=True):
    try:
        def progress(i, total):
            _set_job(progress=i, total=total)

        recs, undated = build_records(io.BytesIO(f1_bytes), date_basis, on_progress=progress)
        if not recs:
            _set_job(status="error", error="No dated rows found in the masterlist.")
            return
        ords = [r[ORD] for r in recs]
        dmin = date.fromordinal(ords[0])
        dmax = date.fromordinal(ords[-1])
        ports = sorted({r[PORT] for r in recs})
        buckets = sorted({r[I2] for r in recs})

        prior = None
        if f2_bytes:
            _set_job(stage="Parsing prior-year file…")
            precs, _ = build_records(io.BytesIO(f2_bytes), date_basis)
            prior = {"records": precs, "ords": [r[ORD] for r in precs]}

        targets_fs = _NamedBytes(targets_bytes, "targets.csv") if targets_bytes else None
        targets = load_targets(targets_fs)

        STATE["ds"] = {"records": recs, "ords": ords, "dmin": dmin, "dmax": dmax,
                       "ports": ports, "buckets": buckets, "targets": targets,
                       "date_basis": date_basis, "file1": f1_name,
                       "undated": undated, "prior": prior, "prior_name": f2_name}
        _set_job(status="done", progress=len(recs), total=len(recs), stage="")

        # Persist to Supabase Storage as a compressed snapshot so the next cold
        # start can restore without re-parsing Excel.  Typically 5–15 MB,
        # well within the 50 MB free-plan limit.  Meta is only written after
        # the snapshot upload succeeds so stale pointers never cause silent
        # reload failures.
        if _save_to_sb and _sb_ok():
            _set_job(stage="Compressing and saving to cloud storage…")
            snap = _pack_ds(STATE["ds"])
            ok = _sb_upload(_SB_SNAP, snap, "application/gzip")
            if ok:
                _sb_put_meta({"file1": f1_name, "file2": f2_name,
                              "date_basis": date_basis,
                              "has_prior": bool(f2_bytes),
                              "has_targets": bool(targets_bytes)})
            else:
                print("[supabase] snapshot upload failed — meta not updated", flush=True)
            _set_job(stage="")
    except Exception:
        _set_job(status="error", error=traceback.format_exc())


@app.route("/upload", methods=["POST"])
def upload():
    if "file1" not in request.files or request.files["file1"].filename == "":
        return "No masterlist file selected.", 400
    date_basis = request.form.get("date_basis", "COLLECTIONDATE")
    if date_basis not in DATE_BASES:
        date_basis = "COLLECTIONDATE"

    f1 = request.files["file1"]
    f1_bytes, f1_name = f1.read(), f1.filename

    f2 = request.files.get("file2")
    f2_bytes = f2.read() if (f2 and f2.filename) else None
    f2_name = f2.filename if f2_bytes else None

    targets_file = request.files.get("targets")
    targets_bytes = targets_file.read() if (targets_file and targets_file.filename) else None

    _set_job(status="parsing", progress=0, total=0, stage="Parsing masterlist…", error=None)
    threading.Thread(target=_run_upload_job,
                      args=(f1_bytes, f1_name, f2_bytes, f2_name, targets_bytes, date_basis),
                      daemon=True).start()
    return render_template("uploading.html")


@app.route("/upload_status")
def upload_status():
    return jsonify(**_get_job())


@app.route("/debug_sb")
def debug_sb():
    meta = _sb_get_meta() if _sb_ok() else None
    return jsonify(
        sb_enabled=_sb_ok(),
        sb_url=bool(_SB_URL),
        sb_key=bool(_SB_KEY),
        autoload_done=_SB_AUTOLOAD_DONE,
        has_data=STATE["ds"] is not None,
        meta=meta,
        job=_get_job(),
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8060))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(debug=debug, host="0.0.0.0", port=port, threaded=True)
